import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook


def excel_to_json():
    wb = load_workbook('excel/佑维.xlsx', data_only=True)
    sheet = wb['books']

    books = []

    for row in islice(sheet.values, 1, sheet.max_row):
        book = {'name': row[0], 'author': row[1], 'intro': row[3]}
        books.append(book)
    return books


if __name__ == '__main__':
    j = excel_to_json()
    with open('data.json', 'w') as f:
        f.write(json.dumps(j, ensure_ascii=False))
